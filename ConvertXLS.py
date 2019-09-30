import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException

def open_xls_as_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()


    for row in range(1, nrows+1):
        for col in range(1, ncols+1):
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row-1, col-1)

    return book1